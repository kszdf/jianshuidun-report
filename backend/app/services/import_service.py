"""
Excel导入服务 - 财务数据导入核心模块
支持：科目余额表、凭证表、序时账
自动识别：表头、期间、科目结构
"""
import os
import uuid
from datetime import datetime, date
from typing import List, Dict, Optional, Tuple
from decimal import Decimal, InvalidOperation
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models.finance import ImpBatch, AccSubjectBalance, AccVoucher, AccVoucherItem
from app.models.config import CfgSubjectMapping
from app.services.mapping_service import auto_map_subjects

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ============================================================
# Excel识别引擎 - 智能识别表头和数据
# ============================================================

def detect_excel_type(file_path: str) -> Tuple[str, int, Dict]:
    """
    智能识别Excel类型和表头行
    返回: (file_type, header_row, column_mapping)
    file_type: balance / voucher / unknown
    """
    if not HAS_OPENPYXL:
        raise Exception("缺少openpyxl依赖，请先安装")

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    # 读取前20行，判断类型
    rows = []
    for i, row in enumerate(ws.iter_rows(max_row=20, values_only=True)):
        rows.append([str(c).strip() if c is not None else "" for c in row])
        if i >= 19:
            break

    wb.close()

    # 识别关键词
    balance_keywords = ["科目编码", "科目代码", "科目名称", "期初余额", "期末余额", "借方", "贷方", "本期发生"]
    voucher_keywords = ["凭证号", "凭证编号", "摘要", "科目", "借方金额", "贷方金额", "记账"]

    balance_score = 0
    voucher_score = 0
    header_row = 0
    column_mapping = {}

    for i, row in enumerate(rows):
        row_text = "".join(row)
        b_score = sum(1 for kw in balance_keywords if kw in row_text)
        v_score = sum(1 for kw in voucher_keywords if kw in row_text)

        if b_score > balance_score and b_score >= 3:
            balance_score = b_score
            header_row = i
            column_mapping = _detect_balance_columns(row)

        if v_score > voucher_score and v_score >= 3:
            voucher_score = v_score
            header_row = i
            column_mapping = _detect_voucher_columns(row)

    if balance_score >= voucher_score and balance_score >= 3:
        return "balance", header_row, column_mapping
    elif voucher_score > balance_score and voucher_score >= 3:
        return "voucher", header_row, column_mapping
    else:
        return "unknown", 0, {}


def _detect_balance_columns(header_row: List[str]) -> Dict:
    """识别科目余额表各列位置"""
    mapping = {}
    for i, cell in enumerate(header_row):
        cell_lower = cell.lower()
        if "科目编码" in cell or "科目代码" in cell or "编码" == cell or "代码" == cell:
            mapping["subject_code"] = i
        elif "科目名称" in cell or "科目" == cell:
            mapping["subject_name"] = i
        elif "科目全称" in cell or "全称" in cell:
            mapping["subject_full_name"] = i
        elif ("期初" in cell and "借" in cell) or ("期初借方" in cell):
            mapping["begin_debit"] = i
        elif ("期初" in cell and "贷" in cell) or ("期初贷方" in cell):
            mapping["begin_credit"] = i
        elif ("本期" in cell and "借" in cell) or ("借方发生" in cell) or "本期借方" in cell:
            mapping["current_debit"] = i
        elif ("本期" in cell and "贷" in cell) or ("贷方发生" in cell) or "本期贷方" in cell:
            mapping["current_credit"] = i
        elif ("期末" in cell and "借" in cell) or ("期末借方" in cell) or ("借" == cell and "期末" in header_row[i-1] if i>0 else False):
            mapping["end_debit"] = i
        elif ("期末" in cell and "贷" in cell) or ("期末贷方" in cell) or ("贷" == cell and "期末" in header_row[i-1] if i>0 else False):
            mapping["end_credit"] = i
        elif ("本年" in cell and "借" in cell) or "本年借方" in cell:
            mapping["year_debit"] = i
        elif ("本年" in cell and "贷" in cell) or "本年贷方" in cell:
            mapping["year_credit"] = i
        elif "项目" in cell and "核算" in cell or "辅助核算" in cell:
            mapping["aux_project"] = i
        elif "客户" in cell:
            mapping["aux_customer"] = i
    return mapping


def _detect_voucher_columns(header_row: List[str]) -> Dict:
    """识别凭证表各列位置"""
    mapping = {}
    for i, cell in enumerate(header_row):
        cell_lower = cell.lower()
        if "凭证号" in cell or "凭证编号" in cell:
            mapping["voucher_no"] = i
        elif "日期" in cell and "凭证" not in cell:
            mapping["voucher_date"] = i
        elif "凭证日期" in cell:
            mapping["voucher_date"] = i
        elif "摘要" in cell:
            mapping["summary"] = i
        elif "科目编码" in cell or "科目代码" in cell:
            mapping["subject_code"] = i
        elif "科目名称" in cell:
            mapping["subject_name"] = i
        elif "借方" in cell and "金额" in cell or cell == "借方金额":
            mapping["debit"] = i
        elif "贷方" in cell and "金额" in cell or cell == "贷方金额":
            mapping["credit"] = i
        elif cell == "借方" or (cell == "借" and "金" in "".join(header_row)):
            mapping["debit"] = i
        elif cell == "贷方" or (cell == "贷" and "金" in "".join(header_row)):
            mapping["credit"] = i
        elif "项目" in cell and "核算" not in cell:
            mapping["aux_project"] = i
        elif "客户" in cell:
            mapping["aux_customer"] = i
        elif "供应商" in cell:
            mapping["aux_supplier"] = i
    return mapping


# ============================================================
# 科目余额表导入
# ============================================================

def import_balance_file(db: Session, company_id: int, file_path: str,
                         source_file: str = "", source_software: str = "",
                         period: str = None) -> Dict:
    """
    导入科目余额表
    1. 智能识别表头
    2. 读取数据
    3. 自动科目映射
    4. 写入数据库
    """
    if not HAS_OPENPYXL:
        raise Exception("缺少openpyxl依赖，请先安装 pip install openpyxl")

    # 1. 识别类型
    file_type, header_row, col_mapping = detect_excel_type(file_path)
    if file_type != "balance":
        # 即使识别为voucher也尝试用balance方式读
        if not col_mapping.get("subject_code") or not col_mapping.get("subject_name"):
            raise Exception("无法识别为科目余额表，请检查文件格式")

    # 2. 读取数据
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    all_rows = []
    subjects = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_row:
            continue
        cells = [c if c is not None else "" for c in row]
        if len(cells) < max(col_mapping.values()) + 1 if col_mapping else 0:
            continue

        code = _get_cell_str(cells, col_mapping.get("subject_code", -1))
        name = _get_cell_str(cells, col_mapping.get("subject_name", -1))
        full_name = _get_cell_str(cells, col_mapping.get("subject_full_name", -1)) or name

        # 跳过空行和合计行
        if not code and not name:
            continue
        if "合计" in name or "总计" in name or "合计" in code:
            continue

        subject = {
            "code": code,
            "name": name,
            "full_name": full_name,
            "begin_debit": _get_cell_decimal(cells, col_mapping.get("begin_debit", -1)),
            "begin_credit": _get_cell_decimal(cells, col_mapping.get("begin_credit", -1)),
            "current_debit": _get_cell_decimal(cells, col_mapping.get("current_debit", -1)),
            "current_credit": _get_cell_decimal(cells, col_mapping.get("current_credit", -1)),
            "end_debit": _get_cell_decimal(cells, col_mapping.get("end_debit", -1)),
            "end_credit": _get_cell_decimal(cells, col_mapping.get("end_credit", -1)),
            "year_debit": _get_cell_decimal(cells, col_mapping.get("year_debit", -1)),
            "year_credit": _get_cell_decimal(cells, col_mapping.get("year_credit", -1)),
            "aux_project": _get_cell_str(cells, col_mapping.get("aux_project", -1)),
            "aux_customer": _get_cell_str(cells, col_mapping.get("aux_customer", -1)),
        }
        subjects.append(subject)

    wb.close()

    if not subjects:
        raise Exception("未读取到有效科目数据")

    # 3. 自动科目映射
    mapping_result = auto_map_subjects(db, company_id, subjects)

    # 4. 创建导入批次
    batch_no = f"BAL{datetime.now().strftime('%Y%m%d%H%M%S')}"
    batch = ImpBatch(
        company_id=company_id,
        batch_no=batch_no,
        import_type="balance",
        source_file=source_file,
        source_software=source_software,
        total_count=len(subjects),
        success_count=len(subjects),
        status="completed",
        quality_score=min(int(mapping_result.get("mapped_count", 0) / len(subjects) * 100), 100),
    )
    db.add(batch)
    db.flush()

    # 5. 写入科目余额数据
    if not period:
        period = date.today().strftime("%Y-%m")

    # 获取映射关系
    mappings = db.query(CfgSubjectMapping).filter(
        CfgSubjectMapping.company_id == company_id
    ).all()
    mapping_by_code = {m.source_code: m for m in mappings}

    # 删除该期间旧数据（同批次替换）
    db.query(AccSubjectBalance).filter(
        AccSubjectBalance.company_id == company_id,
        AccSubjectBalance.period == period,
    ).delete(synchronize_session=False)

    for subj in subjects:
        mapping = mapping_by_code.get(subj["code"])
        bal = AccSubjectBalance(
            company_id=company_id,
            batch_id=batch.id,
            period=period,
            subject_code=subj["code"],
            subject_name=subj["name"],
            subject_full_name=subj["full_name"],
            std_subject_id=mapping.std_subject_id if mapping else None,
            std_subject_code=mapping.std_subject_code if mapping else None,
            begin_debit=subj["begin_debit"],
            begin_credit=subj["begin_credit"],
            current_debit=subj["current_debit"],
            current_credit=subj["current_credit"],
            end_debit=subj["end_debit"],
            end_credit=subj["end_credit"],
            year_debit=subj["year_debit"],
            year_credit=subj["year_credit"],
            aux_project=subj["aux_project"] or None,
            aux_customer=subj["aux_customer"] or None,
            has_aux=bool(subj["aux_project"] or subj["aux_customer"]),
        )
        db.add(bal)

    db.commit()

    return {
        "batch_id": batch.id,
        "batch_no": batch_no,
        "import_type": "balance",
        "period": period,
        "total_count": len(subjects),
        "mapping": mapping_result,
        "message": f"成功导入 {len(subjects)} 条科目余额数据",
    }


# ============================================================
# 凭证表导入
# ============================================================

def import_voucher_file(db: Session, company_id: int, file_path: str,
                         source_file: str = "", source_software: str = "") -> Dict:
    """
    导入凭证表/序时账
    """
    if not HAS_OPENPYXL:
        raise Exception("缺少openpyxl依赖")

    # 1. 识别类型
    file_type, header_row, col_mapping = detect_excel_type(file_path)
    if file_type == "unknown":
        raise Exception("无法识别文件格式，请检查是否为凭证表或序时账")

    # 2. 读取凭证数据
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    voucher_items = []
    current_voucher_no = ""
    current_voucher_date = None
    current_summary = ""

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_row:
            continue
        cells = [c if c is not None else "" for c in row]

        voucher_no = _get_cell_str(cells, col_mapping.get("voucher_no", -1))
        voucher_date = _get_cell_date(cells, col_mapping.get("voucher_date", -1))
        summary = _get_cell_str(cells, col_mapping.get("summary", -1))
        subject_code = _get_cell_str(cells, col_mapping.get("subject_code", -1))
        subject_name = _get_cell_str(cells, col_mapping.get("subject_name", -1))
        debit = _get_cell_decimal(cells, col_mapping.get("debit", -1))
        credit = _get_cell_decimal(cells, col_mapping.get("credit", -1))
        aux_project = _get_cell_str(cells, col_mapping.get("aux_project", -1))
        aux_customer = _get_cell_str(cells, col_mapping.get("aux_customer", -1))
        aux_supplier = _get_cell_str(cells, col_mapping.get("aux_supplier", -1))

        # 跳过空行
        if not subject_code and not subject_name and debit == 0 and credit == 0:
            continue

        # 空凭证号沿用上面的
        if voucher_no:
            current_voucher_no = voucher_no
        if voucher_date:
            current_voucher_date = voucher_date
        if summary:
            current_summary = summary

        if not subject_code and not subject_name:
            continue

        voucher_items.append({
            "voucher_no": current_voucher_no,
            "voucher_date": current_voucher_date,
            "summary": summary or current_summary,
            "subject_code": subject_code,
            "subject_name": subject_name,
            "debit": debit,
            "credit": credit,
            "aux_project": aux_project,
            "aux_customer": aux_customer,
            "aux_supplier": aux_supplier,
        })

    wb.close()

    if not voucher_items:
        raise Exception("未读取到有效凭证数据")

    # 3. 提取唯一科目列表，做自动映射
    unique_subjects = []
    seen_codes = set()
    for item in voucher_items:
        code = item["subject_code"]
        if code and code not in seen_codes:
            seen_codes.add(code)
            unique_subjects.append({
                "code": code,
                "name": item["subject_name"],
                "full_name": item["subject_name"],
            })

    mapping_result = auto_map_subjects(db, company_id, unique_subjects)

    # 4. 创建导入批次
    batch_no = f"VOU{datetime.now().strftime('%Y%m%d%H%M%S')}"

    # 计算期间范围
    dates = [item["voucher_date"] for item in voucher_items if item["voucher_date"]]
    period_start = min(dates).strftime("%Y-%m-%d") if dates else None
    period_end = max(dates).strftime("%Y-%m-%d") if dates else None

    batch = ImpBatch(
        company_id=company_id,
        batch_no=batch_no,
        import_type="voucher",
        source_file=source_file,
        source_software=source_software,
        total_count=len(voucher_items),
        success_count=len(voucher_items),
        status="completed",
        quality_score=min(int(mapping_result.get("mapped_count", 0) / max(len(unique_subjects), 1) * 100), 100),
    )
    db.add(batch)
    db.flush()

    # 5. 写入凭证和凭证明细
    # 获取映射关系
    mappings = db.query(CfgSubjectMapping).filter(
        CfgSubjectMapping.company_id == company_id
    ).all()
    mapping_by_code = {m.source_code: m for m in mappings}

    # 按凭证号分组
    voucher_groups = {}
    for idx, item in enumerate(voucher_items):
        key = f"{item['voucher_no']}_{item['voucher_date']}" if item['voucher_date'] else item["voucher_no"]
        if key not in voucher_groups:
            voucher_groups[key] = []
        voucher_groups[key].append((idx, item))

    # 写入凭证主表和明细表
    voucher_count = 0
    for key, items in voucher_groups.items():
        first_item = items[0][1]
        total_debit = sum(item[1]["debit"] for item in items)
        total_credit = sum(item[1]["credit"] for item in items)

        period = first_item["voucher_date"].strftime("%Y-%m") if first_item["voucher_date"] else ""

        voucher = AccVoucher(
            company_id=company_id,
            batch_id=batch.id,
            voucher_no=first_item["voucher_no"],
            voucher_date=first_item["voucher_date"],
            period=period,
            summary=first_item["summary"],
            total_debit=total_debit,
            total_credit=total_credit,
        )
        db.add(voucher)
        db.flush()
        voucher_count += 1

        for line_no, (idx, item) in enumerate(items, 1):
            mapping = mapping_by_code.get(item["subject_code"])
            v_item = AccVoucherItem(
                company_id=company_id,
                voucher_id=voucher.id,
                voucher_date=item["voucher_date"],
                period=period,
                line_no=line_no,
                subject_code=item["subject_code"],
                subject_name=item["subject_name"],
                std_subject_id=mapping.std_subject_id if mapping else None,
                std_subject_code=mapping.std_subject_code if mapping else None,
                summary=item["summary"],
                debit=item["debit"],
                credit=item["credit"],
                aux_project=item["aux_project"] or None,
                aux_customer=item["aux_customer"] or None,
                aux_supplier=item["aux_supplier"] or None,
            )
            db.add(v_item)

    db.commit()

    return {
        "batch_id": batch.id,
        "batch_no": batch_no,
        "import_type": "voucher",
        "period_start": period_start,
        "period_end": period_end,
        "total_items": len(voucher_items),
        "voucher_count": voucher_count,
        "subject_count": len(unique_subjects),
        "mapping": mapping_result,
        "message": f"成功导入 {voucher_count} 张凭证，共 {len(voucher_items)} 条分录",
    }


# ============================================================
# 工具函数
# ============================================================

def _get_cell_str(cells: List, index: int) -> str:
    """安全获取单元格字符串"""
    if index < 0 or index >= len(cells):
        return ""
    val = cells[index]
    if val is None:
        return ""
    return str(val).strip()


def _get_cell_decimal(cells: List, index: int) -> Decimal:
    """安全获取单元格数值"""
    if index < 0 or index >= len(cells):
        return Decimal("0")
    val = cells[index]
    if val is None or val == "":
        return Decimal("0")
    try:
        if isinstance(val, (int, float)):
            return Decimal(str(val))
        return Decimal(str(val).replace(",", "").replace("，", ""))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _get_cell_date(cells: List, index: int) -> Optional[date]:
    """安全获取单元格日期"""
    if index < 0 or index >= len(cells):
        return None
    val = cells[index]
    if val is None or val == "":
        return None
    if isinstance(val, date):
        return val
    if isinstance(val, datetime):
        return val.date()
    # 尝试解析字符串
    val_str = str(val).strip()
    for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日", "%m/%d/%Y"]:
        try:
            return datetime.strptime(val_str, fmt).date()
        except ValueError:
            continue
    return None


def save_upload_file(file_content: bytes, filename: str) -> str:
    """保存上传文件到本地"""
    os.makedirs(settings.UPLOAD_DIR, exist_ok=True)
    ext = os.path.splitext(filename)[1] or ".xlsx"
    unique_name = f"{uuid.uuid4().hex}{ext}"
    file_path = os.path.join(settings.UPLOAD_DIR, unique_name)
    with open(file_path, "wb") as f:
        f.write(file_content)
    return file_path
